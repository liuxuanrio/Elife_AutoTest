from openpyxl.reader.excel import load_workbook

from config.projectConfig import ProjectConfig
from utils.config import FileDate, TimeMethod, DataType
import traceback
from test.common.web_Driver_run import WebDriverRun


class AutoFile(WebDriverRun):

    """
    运行mqt脚本
    """
    def runMqtFile(self, filename, caseutc):
        # 获取工作目录
        script_path = FileDate().osFilePath()

        # 拼接mqt脚本文件路径
        picturePath = f"{script_path}/data/test_picture/{filename}"

        # 保存时间和路径
        self.caseutc = caseutc
        self.path = picturePath

        # 判断是否在jenkins运行
        if "jenkins_home" in picturePath:
            self.system_test = 1

        # 获取配置文件中测试数据路径
        configData = OpenFile().runConfig(script_path)

        # 保存测试文件路径
        self.testData = configData["testData"]

        # 保存脚本路径
        mqtPath = f'{configData["testCase"]}{filename}'

        # 获取脚本文件list数据
        mqtList = self.openFile(mqtPath)
        print(mqtList)

        # 调用执行方法
        for mqt in mqtList:
            self.transfer(mqt[0], mqt[1])

        # 全局变量保存到log
        self.runlog += f"\n 打开的浏览器：{str(self.chrstrlist)}" \
                       f"\n 打开的浏览器list：{str(self.chrlist)}" \
                       f"\n 保存的变量信息：{str(self.globalVariable)}" \
                       f"\n 保存的判断结果：{str(self.ifstat)}"
        self.logs(self.assertLog)

        assertint = self.assertLog
        runlogData = self.runlog
        variable = self.globalVariable

        # 关闭开启的浏览器
        self.webquit(picturePath, caseutc)

        return runlogData, assertint, variable

    """
    读取mqt文件, 生成list
    """
    def openFile(self, mqtPath):
        # 定义mqtList保存mqt脚本处理后的list
        mqtList = []

        # 读取文件
        file = open(mqtPath, 'r', encoding='utf8')

        # 读取文件为list
        lists = file.readlines()[2:-1]

        # 定义tuplesum 计算每一行为(就加1，为)就减1
        tuplesum = 0

        # 定义valuelist 保存每个步骤
        valuelist = []

        # 定义key 保存步骤执行的方法
        key = ""

        # 循环读取文件每行内容
        for i in range(len(lists)):
            # 清除每一行的头尾的空格
            data = lists[i].strip()

            # 判断 (tuplesum 等于0 and 当前行内容不等于( and 不是注释) or 当前是最后一行
            if (tuplesum == 0 and data != "(" and "^" not in data) or i == len(lists) - 1:
                # 当前是最后一行
                if i == len(lists) - 1:
                    # 把最后一行内容加到valuelist
                    valuelist.append(data)
                if len(valuelist) > 0:  # 判断value是否不为空
                    # 判断valuelist 第一个值等于'('
                    if valuelist[0] == "(":
                        pass
                        # 调用self.for_list方法剥离嵌套方法，去除list前后的括号
                        valuelist = self.for_list(valuelist[1:-1])
                    # 保存步骤信息
                    mqtList.append([key, valuelist])
                    print(key, valuelist)
                # 执行方法运行后，value设置为空
                valuelist = []
                # 保存符合条件的字符为key
                key = data
            else:
                if data == "(":
                    tuplesum += 1
                    valuelist.append(data)
                elif data == ")":
                    tuplesum -= 1
                    valuelist.append(data)
                else:
                    if "^" in data:
                        pass
                    else:
                        valuelist.append(DataType().updateStrMake(data))
        # 关闭文件
        file.close()

        return mqtList

    # 处理list中嵌套方法
    def for_list(self, filelist):
        if "(" in filelist:
            tuplesum = 0
            valuelist = []
            key = ""
            return_list = []
            for i in range(len(filelist)):
                data = filelist[i]
                if (tuplesum == 0 and data != "(" and "^" not in data) or i == len(filelist) - 1:
                    if i == len(filelist) - 1:
                        valuelist.append(data)
                    if len(key) > 0:  # 判断value是否不为空
                        return_list.append(key)
                    if len(valuelist) > 0:
                        if valuelist[0] == "(":
                            valuelist = self.for_list(valuelist[1:-1])
                        elif type(valuelist) == list:
                            valuelist = valuelist[0]
                        return_list.append(valuelist)
                    # 执行方法运行后，value设置为空
                    valuelist = []
                    # 保存符合条件的字符为key
                    key = DataType().updateStrMake(data)
                else:
                    if data == "(":
                        tuplesum += 1
                        valuelist.append(data)
                    elif data == ")":
                        tuplesum -= 1
                        valuelist.append(data)
                    else:
                        if "^" in data:
                            self.logs(data)
                        else:
                            if data == "', '":
                                pass
                                valuelist.append(data)
                            else:
                                valuelist.append(DataType().updateStrMake(data))
            return return_list
        else:
            return filelist

    # 处理后的值调用对应的方法
    def transfer(self, key, value):
        try:
            self.main(key, value)
        except:
            errorlog = traceback.print_exc()
            self.logs(f"执行失败：{errorlog}-------------------------------------")


class OpenFile():
    # 读取用例表格并转化成list
    def openXlsx(self, path, fileName, module):
        excelDir = fr'{path}/test/case/{fileName}'  # 打开xlsx文件
        workBook = load_workbook(excelDir)  # 保存原样--样式
        # 2- 操作对应的用例表
        workSheet = workBook.worksheets[0]
        dataList = []
        cnt = 1
        if module:
            for cnts in workSheet.rows:
                if cnt > 2:
                    caseProject = workSheet.cell(cnt, 1).value  # 用例项目
                    caseModule = workSheet.cell(cnt, 2).value  # 用例模块
                    caseTitle = workSheet.cell(cnt, 3).value  # 用例标题
                    caseDescription = workSheet.cell(cnt, 4).value   # 用例描述
                    casePoint = workSheet.cell(cnt, 5).value   # 测试点
                    if module in caseModule:
                        dataList.append([caseProject, caseModule, caseTitle, caseDescription, casePoint])
                cnt += 1
        else:
            for cnts in workSheet.rows:
                if cnt > 2:
                    caseProject = workSheet.cell(cnt, 1).value  # 用例项目
                    caseModule = workSheet.cell(cnt, 2).value  # 用例模块
                    caseTitle = workSheet.cell(cnt, 3).value  # 用例标题
                    caseDescription = workSheet.cell(cnt, 4).value   # 用例描述
                    casePoint = workSheet.cell(cnt, 5).value   # 测试点
                    dataList.append([caseProject, caseModule, caseTitle, caseDescription, casePoint])
                cnt += 1
        return dataList  # 列表

    # 处理单独运行的脚本数据
    def updateFileList(self, projectName, fileList):
        dataList = []
        for name in fileList:
            dataList.append([projectName, projectName, name, projectName, projectName])
        return dataList

    # 获取执行脚本list
    def testFileCase(self):
        path = FileDate().osFilePath()  # 获取文件路径
        configData = self.runConfig(path)
        runType = configData["runType"]
        fileName = configData["fileName"]
        if runType:
            if "," in runType:
                file_name_list = runType.split(",")
            else:
                file_name_list = [runType]
            file_name_list = self.updateFileList("driver_app", file_name_list)
        else:
            file_name_list = self.openXlsx(path, fileName, "")
        return file_name_list

    def testCaseAll(self, module):
        path = FileDate().osFilePath()  # 获取文件路径
        configData = self.runConfig(path)
        fileName = configData["fileName"]
        file_name_list = self.openXlsx(path, fileName, module)
        return file_name_list

    def runConfig(self, path):
        fileConfig = ProjectConfig().projectConfig()
        # 判断当前项目运行环境
        if "jenkins_home" in path:
            if "Elife_UI_AutoTest" in path:
                keys = "driver_app"
            elif "Flt_mgmt_UI_AutoTest" in path:
                keys = "flt_mgmt"
            else:
                keys = "driver_app"
        else:  # 本地调试可手动修改需要执行的文件
            keys = "driver_app"
        fileName = fileConfig[keys]["caseName"]
        runType = fileConfig[keys]["runType"]
        testData = f'{path}{fileConfig[keys]["testData"]}'
        testCase = f'{path}{fileConfig[keys]["testCase"]}'
        jobName = fileConfig[keys]["jobName"]
        title = fileConfig[keys]["title"]
        configData = {"fileName": fileName, "runType": runType,
                      "testData": testData, "testCase": testCase,
                      "jobName": jobName, "title": title}
        return configData



if __name__ == "__main__":
    pass
    caseutc = TimeMethod().intNewTimeUtc()
    case = AutoFile().runMqtFile("driver_app_account_add.mqt", caseutc)
    print(case)
    # case = print(OpenFile().testFileCase())